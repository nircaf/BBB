import pandas as pd
import os
import numpy as np

def read_all_excels(directory='Volume'):
    def get_excels(sheet_name):
        dataframes = {}
        df_p = pd.DataFrame()
        for excel_file in excel_files:
            df = pd.read_excel(os.path.join(directory, excel_file), sheet_name=sheet_name)
            dataframes[excel_file] = df
            df2 = pd.DataFrame()
            df2[excel_file.strip('.xlsx')] = df.drop(columns=['ID']).mean()
            df_p = pd.concat([df_p, df2], axis=1)
        return df_p
    # Get all files in the directory
    files = os.listdir(directory)
    # Filter for Excel files
    excel_files = [f for f in files if f.endswith('.xlsx') or f.endswith('.xls')]
    # Read each Excel file into a DataFrame and store them in a dictionary
    df_p = get_excels('%')
    create_scientific_boxplot(df_p,y_label="%, Volume",filename = 'Volume')
    df_p = get_excels('Cm3')
    create_scientific_boxplot(df_p,y_label="Cm3",filename = 'Cm3')
    return df_p



def get_folders():
    this_folder = os.getcwd()
    files_location = this_folder + r"\Excel not age gender match"
    excel_file = files_location + r"\epilepsy\merged.xlsx"
    controls_file = files_location + r"\controls\merged.xlsx"
    clinical_data = this_folder + r"\Epilepsy_clinical_data.xlsx"
    return files_location, excel_file, controls_file, clinical_data

import stats

def get_results():
    # this folder
    files_location, excel_file, controls_file, clinical_data = get_folders()
    clinical_data_df = pd.read_excel(clinical_data, sheet_name="All")
    # Epilepsy_clinical_data drop after row 47
    nan_code_index = clinical_data_df[clinical_data_df["code"].isna()].index[0]
    clinical_data_df = clinical_data_df.iloc[:nan_code_index]
    Epilepsy_clinical_data = clinical_data_df
    di = {}
    result_mat_lin_age = pd.read_excel(excel_file, sheet_name="BBB_percent_Linear")
    result_mat_lin_age = get_126_areas(clinical_data_df, result_mat_lin_age)
    # rename Lin to y_target
    result_mat_lin_age = result_mat_lin_age.rename(columns={"Lin": "y_target"})
    controls_age_gender = pd.read_excel(clinical_data, sheet_name="Controls")
    controls_lin = pd.read_excel(controls_file, sheet_name="BBB_percent_Linear")
    #  (mean age = 27.39 ± 9.98, 53.57% male)
    # get age mean and std
    di["age_mean"] = Epilepsy_clinical_data["'age'"].mean()
    di["age_std"] = Epilepsy_clinical_data["'age'"].std()
    epilepsy_gender = Epilepsy_clinical_data["'gender'"]
    epilepsy_gender = epilepsy_gender.str.strip("'").str.strip().str.upper()
    # Epilepsy_clinical_data["'gender'"]=="'M'" or Epilepsy_clinical_data["'gender'"]=='M'
    di["male_mean"] = (sum(epilepsy_gender == "M")) / len(epilepsy_gender)
    di["BBBP_lin_mean"] = result_mat_lin_age["y_target"].mean()
    di["BBBP_lin_std"] = result_mat_lin_age["y_target"].std()
    control_gender = controls_age_gender["Gender"]
    controls_age = controls_age_gender["age"]
    control_gender = control_gender.str.strip("'").str.strip().str.upper()
    di["male_mean_controls"] = (sum(control_gender == "M")) / len(control_gender)
    di[
        "str_age_gender"
    ] = f'mean age = {di["age_mean"]:.2f} ± {di["age_std"]:.2f}, {di["male_mean"]*100:.2f}% male'
    focal_epilepsy_age = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal/General"] == "F"
    ]["'age'"]
    focal_epilepsy_gender = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal/General"] == "F"
    ]["'gender'"]
    focal_epilepsy_gender = focal_epilepsy_gender.str.strip("'").str.strip().str.upper()
    male_mean_f = (sum(focal_epilepsy_gender == "M")) / len(focal_epilepsy_gender)
    di[
        "str_age_focal"
    ] = f"mean age = {focal_epilepsy_age.mean():.2f} ± {focal_epilepsy_age.std():.2f}, {male_mean_f*100:.2f}% male"
    g_epilepsy_age = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal/General"] == "G"
    ]["'age'"]
    g_epilepsy_gender = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal/General"] == "G"
    ]["'gender'"]
    g_epilepsy_gender = g_epilepsy_gender.str.strip("'").str.strip().str.upper()
    male_mean_g = (sum(g_epilepsy_gender == "M")) / len(g_epilepsy_gender)
    di[
        "str_age_g"
    ] = f"mean age = {g_epilepsy_age.mean():.2f} ± {g_epilepsy_age.std():.2f}, {male_mean_g*100:.2f}% male"
    di[
        "str_age_gender_controls"
    ] = f'mean age = {controls_age.mean():.2f} ± {controls_age.std():.2f}, {di["male_mean_controls"]*100:.2f}% male'
    # do mann whitney u test on age
    from scipy.stats import mannwhitneyu

    stat, di["p_age_controls_epilepsy"] = mannwhitneyu(
        controls_age, Epilepsy_clinical_data["'age'"]
    )
    # mann whitney on gender
    stat, di["p_gender_controls_epilepsy"] = mannwhitneyu(
        control_gender == "M", Epilepsy_clinical_data["'gender'"].dropna() == "M"
    )
    # mann whitney on focal epilepsy age
    stat, di["p_age_focal_general"] = mannwhitneyu(controls_age, focal_epilepsy_age)
    # mann whitney on focal epilepsy gender
    g_epilepsy_gender = g_epilepsy_gender.dropna().reset_index(drop=True)
    focal_epilepsy_gender = focal_epilepsy_gender.dropna().reset_index(drop=True)
    stat, di["p_gender_focal_general"] = mannwhitneyu(
        g_epilepsy_gender == "M", focal_epilepsy_gender == "M"
    )
    frontal_epilepsy = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal Type"].str.contains("F") == True
    ]
    frontal_epilepsy_age = frontal_epilepsy["'age'"]
    frontal_epilepsy_gender = frontal_epilepsy["'gender'"]
    frontal_epilepsy_gender = (
        frontal_epilepsy_gender.str.strip("'").str.strip().str.upper()
    )
    male_mean_frontal = (sum(frontal_epilepsy_gender == "M")) / len(
        frontal_epilepsy_gender
    )
    temporal_epilepsy = Epilepsy_clinical_data[
        Epilepsy_clinical_data["Focal Type"].str.contains("T") == True
    ]
    temporal_epilepsy_age = temporal_epilepsy["'age'"]
    temporal_epilepsy_gender = temporal_epilepsy["'gender'"]
    temporal_epilepsy_gender = (
        temporal_epilepsy_gender.str.strip("'").str.strip().str.upper()
    )
    male_mean_temporal = (sum(temporal_epilepsy_gender == "M")) / len(
        temporal_epilepsy_gender
    )
    stat, di["p_age_frontal_temporal"] = mannwhitneyu(
        frontal_epilepsy_age, temporal_epilepsy_age
    )
    stat, di["p_gender_frontal_temporal"] = mannwhitneyu(
        frontal_epilepsy_gender == "M", temporal_epilepsy_gender == "M"
    )
    di[
        "paper"
    ] = f"""To conduct the study we included {len(Epilepsy_clinical_data)} epileptic patients (mean age = {di["age_mean"]:.2f} ±
    {di["age_std"]:.2f}, {di['male_mean']*100:.2f}% male).
    And {len(controls_lin)} healthy controls (mean age = {controls_age.mean():.2f} ± {controls_age.std():.2f}, {100*di['male_mean_controls']:.2f}% male).
    Among the epileptic patients, {len(focal_epilepsy_age)} had focal epilepsy
    (mean age = {focal_epilepsy_age.mean():.2f} ± {focal_epilepsy_age.std():.2f}, {male_mean_f*100:.2f}% male),
    this group divides into {len(frontal_epilepsy)} frontal epileptic patients (mean age = {frontal_epilepsy_age.mean():.2f} ± {frontal_epilepsy_age.std():.2f}, {male_mean_frontal*100:.2f}% male),
    and to {len(temporal_epilepsy)} temporal epileptic patients (mean age = {temporal_epilepsy_age.mean():.2f} ± {temporal_epilepsy_age.std():.2f}, {male_mean_temporal*100:.2f}% male).
    and {len(g_epilepsy_age)} had generalized epilepsy
    (mean age = {g_epilepsy_age.mean():.2f} ± {g_epilepsy_age.std():.2f}, {male_mean_g*100:.2f}% male).
    After conducting Mann-Whitney U tests, there were no significant differences in age (p={di['p_age_controls_epilepsy']:.2f})
    or gender distribution (p={di['p_gender_controls_epilepsy']:.2f}) between the controls and epileptic patients. Additionally,
    there were no significant differences between focal and generalized epileptic patients in age (p={di['p_age_focal_general']:.2f})
    or gender distribution (p={di['p_gender_focal_general']:.2f}).
    Finally, there were no significant differences between frontal and temporal epileptic patients in age (p={di['p_age_frontal_temporal']:.2f}) or gender distribution (p={di['p_gender_frontal_temporal']:.2f}).
    """.replace(
        "\n", " "
    )
    import scipy.stats as stats

    # stip the "'" from columns
    Epilepsy_clinical_data.columns = Epilepsy_clinical_data.columns.str.strip("'")
    # df. index =  Age, Mean, Age at onset median, years (range) Median (IQR), Epilepsy duration , median, years (range),Polytherapy. columns = Controls, Epilepsy, Focal Epilepsy, Generalized Epilepsy, Temporal Epilepsy, df	X2	P
    df_table = pd.DataFrame(index=["Age (years)", "Age at onset median, years (range)", "Epilepsy duration , median, years (range)"], columns=["Controls", "Epilepsy", "Focal Epilepsy", "Generalized Epilepsy", "Temporal Epilepsy", "Frontal Epilepsy","df","X2","P"])
    df_table.loc['Age (years)', 'Controls'] = f"{controls_age.mean():.2f} ({controls_age.std():.2f})"
    df_table.loc['Age (years)', 'Epilepsy'] = f"{Epilepsy_clinical_data['age'].mean():.2f} ({Epilepsy_clinical_data['age'].std():.2f})"
    df_table.loc['Age (years)', 'Focal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F')==True]['age'].mean():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F')==True]['age'].std():.2f})"
    df_table.loc['Age (years)', 'Generalized Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G')==True]['age'].mean():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G')==True]['age'].std():.2f})"
    df_table.loc['Age (years)', 'Temporal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T')==True]['age'].mean():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T')==True]['age'].std():.2f})"
    df_age = pd.DataFrame(
        {
            'Controls': controls_age,
            'Epilepsy': Epilepsy_clinical_data['age'],
            'Focal Epilepsy': Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F')==True]['age'],
            'Generalized Epilepsy': Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G')==True]['age'],
            'Temporal Epilepsy': Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T')==True]['age'],
        }
    )
    # do ANOVA on df_age
    f_val, p_val = stats.f_oneway(df_age['Controls'].dropna(), df_age['Epilepsy'].dropna(), df_age['Focal Epilepsy'].dropna(), df_age['Generalized Epilepsy'].dropna(), df_age['Temporal Epilepsy'].dropna())
    df_table.loc['Age (years)', 'df'] = 3
    df_table.loc['Age (years)', 'F'] = f"{f_val:.2f}"
    df_table.loc['Age (years)', 'P'] = f"{p_val:.2f}"
    df_table.loc['Gender, female %', 'Controls'] = f"{sum(control_gender=='F')/len(control_gender)*100:.2f}"
    df_table.loc['Gender, female %', 'Epilepsy'] = f"{sum(epilepsy_gender=='F')/len(epilepsy_gender)*100:.2f}"
    df_table.loc['Gender, female %', 'Focal Epilepsy'] = f"{epilepsy_gender[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True].value_counts()['F']/len(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True])*100:.2f}"
    df_table.loc['Gender, female %', 'Generalized Epilepsy'] = f"{epilepsy_gender[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True].value_counts()['F']/len(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True])*100:.2f}"
    df_table.loc['Gender, female %', 'Temporal Epilepsy'] = f"{epilepsy_gender[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True].value_counts()['F']/len(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True])*100:.2f}"
    df_gender = pd.DataFrame({'Controls': (control_gender=='M').astype(int),
                            'Epilepsy': (epilepsy_gender=='M').astype(int),
                            'Focal Epilepsy': (epilepsy_gender[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]=='M').astype(int),
                            'Generalized Epilepsy': (epilepsy_gender[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]=='M').astype(int),
                            'Temporal Epilepsy': (epilepsy_gender[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]=='M').astype(int),
                            })
    # Assuming df_gender is your DataFrame and 'Controls', 'Epilepsy', 'Focal Epilepsy', 'Generalized Epilepsy', 'Temporal Epilepsy' are your column names
    observed = pd.concat([df_gender['Controls'].dropna().reset_index(drop=True), df_gender['Epilepsy'].dropna().reset_index(drop=True), df_gender['Focal Epilepsy'].dropna().reset_index(drop=True), df_gender['Generalized Epilepsy'].dropna().reset_index(drop=True), df_gender['Temporal Epilepsy'].dropna().reset_index(drop=True)], axis=1)
    # Perform the Chi-Square test
    chi2, p, dof, expected = stats.chi2_contingency(observed)
    df_table.loc['Age at onset median, years (range)', 'Epilepsy'] = f"{Epilepsy_clinical_data['Age of oneset'].median():.2f} ({Epilepsy_clinical_data['Age of oneset'].min():.2f}-{Epilepsy_clinical_data['Age of oneset'].max():.2f})"
    df_table.loc['Epilepsy duration , median, years (range)', 'Epilepsy'] = f"{Epilepsy_clinical_data['Year of epilepsy'].median():.2f} ({Epilepsy_clinical_data['Year of epilepsy'].min():.2f}-{Epilepsy_clinical_data['Year of epilepsy'].max():.2f})"
    df_table.loc['Age at onset median, years (range)', 'Focal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Age of oneset'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Age of oneset'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Age of oneset'].max():.2f})"
    df_table.loc['Epilepsy duration , median, years (range)', 'Focal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Year of epilepsy'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Year of epilepsy'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Year of epilepsy'].max():.2f})"
    df_table.loc['Age at onset median, years (range)', 'Generalized Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Age of oneset'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Age of oneset'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Age of oneset'].max():.2f})"
    df_table.loc['Epilepsy duration , median, years (range)', 'Generalized Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Year of epilepsy'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Year of epilepsy'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Year of epilepsy'].max():.2f})"
    df_table.loc['Age at onset median, years (range)', 'Temporal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Age of oneset'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Age of oneset'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Age of oneset'].max():.2f})"
    df_table.loc['Epilepsy duration , median, years (range)', 'Temporal Epilepsy'] = f"{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Year of epilepsy'].median():.2f} ({Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Year of epilepsy'].min():.2f}-{Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Year of epilepsy'].max():.2f})"
    f_val, p_val = stats.f_oneway(Epilepsy_clinical_data['Age of oneset'].dropna(),
                                   Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Age of oneset'].dropna()
                                   ,Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Age of oneset'].dropna()
                                   ,Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Age of oneset'].dropna())
    df_table.loc['Age at onset median, years (range)','df'] = 3
    df_table.loc['Age at onset median, years (range)','F'] = f_val
    df_table.loc['Age at onset median, years (range)','P'] = p_val
    f_val, p_val = stats.f_oneway(Epilepsy_clinical_data['Year of epilepsy'].dropna(),
                                   Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Year of epilepsy'].dropna()
                                   ,Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Year of epilepsy'].dropna()
                                   ,Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Year of epilepsy'].dropna())
    df_table.loc['Epilepsy duration , median, years (range)','df'] = 3
    df_table.loc['Epilepsy duration , median, years (range)','F'] = f_val
    df_table.loc['Epilepsy duration , median, years (range)','P'] = p_val

    df_table.loc['Lesional %', 'Epilepsy'] = 100*(Epilepsy_clinical_data['Lesion'] != 0).mean()
    df_table.loc['Lesional %', 'Focal Epilepsy'] = 100*(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('F') == True]['Lesion'] != 0).mean()
    df_table.loc['Lesional %', 'Generalized Epilepsy'] = 100*(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].str.contains('G') == True]['Lesion'] != 0).mean()
    df_table.loc['Lesional %', 'Temporal Epilepsy'] = 100*(Epilepsy_clinical_data[Epilepsy_clinical_data['Focal Type'].str.contains('T') == True]['Lesion'] != 0).mean()
    # create df of result_mat_lin_age["y_target"] and controls_lin["Linear"]
    df_BBB_percent = pd.DataFrame(
        {
            "Controls": controls_lin["Lin"],
            "Epilepsy": result_mat_lin_age["y_target"],
            "Focal Epilepsy": result_mat_lin_age[
                result_mat_lin_age["ID"]
                .iloc[:, 1]
                .isin(
                    clinical_data_df[
                        clinical_data_df["Focal/General"].str.contains("F") == True
                    ]["code"]
                )
            ]["y_target"].reset_index(drop=True),
            "Generalized Epilepsy": result_mat_lin_age[
                result_mat_lin_age["ID"]
                .iloc[:, 1]
                .isin(
                    clinical_data_df[
                        clinical_data_df["Focal/General"].str.contains("G") == True
                    ]["code"]
                )
            ]["y_target"].reset_index(drop=True),
            "Frontal Epilepsy": result_mat_lin_age[
                result_mat_lin_age["ID"]
                .iloc[:, 1]
                .isin(
                    clinical_data_df[
                        clinical_data_df["Focal Type"].str.contains("F") == True
                    ]["code"]
                )
            ]["y_target"].reset_index(drop=True),
            "Temporal Epilepsy": result_mat_lin_age[
                result_mat_lin_age["ID"]
                .iloc[:, 1]
                .isin(
                    clinical_data_df[
                        clinical_data_df["Focal Type"].str.contains("T") == True
                    ]["code"]
                )
            ]["y_target"].reset_index(drop=True),
        }
    )
    result_mat_lin_age126 = pd.read_excel(excel_file, sheet_name="126_Regions_Linear")
    controls_126 = pd.read_excel(controls_file, sheet_name="126_Regions_Linear").drop(
        columns=["'ID'"]
    )
    controls_126.columns = controls_126.columns.str.strip("'")
    controls_126_mean = controls_126.median()
    controls_126_std = controls_126.apply(lambda x: np.mean(np.abs(x - controls_126_mean[x.name])))
    df_126 = get_126_areas(clinical_data_df, result_mat_lin_age126)
    df_126_all = df_126.drop(columns=["ID"])
    df_percent_126 = pd.DataFrame({'ID': df_126['ID'], 'Epilepsy': df_126_all.apply(
                lambda row: 100
                * sum((row - controls_126_mean) / controls_126_std >= 2)
                / len(controls_126_mean),
                axis=1,
            ).reset_index(drop=True)})
    df_126_focal = (
        df_126[
            df_126["ID"].isin(
                clinical_data_df[
                    clinical_data_df["Focal/General"].str.contains("F") == True
                ]["code"]
            )
        ]
        .reset_index(drop=True)
        .drop(columns=["ID"])
    )
    df_126_general = (
        df_126[
            df_126["ID"].isin(
                clinical_data_df[
                    clinical_data_df["Focal/General"].str.contains("G") == True
                ]["code"]
            )
        ]
        .reset_index(drop=True)
        .drop(columns=["ID"])
    )
    df_126_frontal = (
        df_126[
            df_126["ID"].isin(
                clinical_data_df[
                    clinical_data_df["Focal Type"].str.contains("F") == True
                ]["code"]
            )
        ]
        .reset_index(drop=True)
        .drop(columns=["ID"])
    )
    # do mannwhitneyu test on each column general vs df_126_focal
    df_manu = pd.DataFrame()
    for col in df_126_general.columns:
        # Convert the columns to numeric type before dropping NA values and running the test
        df_126_general[col] = pd.to_numeric(df_126_general[col], errors='coerce')
        df_126_focal[col] = pd.to_numeric(df_126_focal[col], errors='coerce')
        df_126_all[col] = pd.to_numeric(df_126_all[col], errors='coerce')
        controls_126[col] = pd.to_numeric(controls_126[col], errors='coerce')
        try:
            # print(mannwhitneyu(df_126_general[col].dropna(), df_126_focal[col].dropna()))
            # df_manu[col] = mannwhitneyu(df_126_general[col].dropna(), df_126_focal[col].dropna())
            df_manu[col] = mannwhitneyu(df_126_all[col].dropna(), controls_126[col].dropna())
        except:
            pass
    # print sum df_manu row 1 is less than 0.05
    print(sum(df_manu.iloc[1,:]< 0.05))
    df_126_temporal = (
        df_126[
            df_126["ID"].isin(
                clinical_data_df[
                    clinical_data_df["Focal Type"].str.contains("T") == True
                ]["code"]
            )
        ]
        .reset_index(drop=True)
        .drop(columns=["ID"])
    )
    df_126 = df_126.drop(columns=["ID"])
    # for each row in result_mat_lin_age get % of areas above 2 sd of controls
    df_2bbbd = pd.DataFrame(
        {
            "Controls": controls_126.apply(
                lambda row:
                sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
            "Epilepsy": df_126.apply(
                lambda row:  sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
            "Focal Epilepsy": df_126_focal.apply(
                lambda row:  sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
            "Generalized Epilepsy": df_126_general.apply(
                lambda row:  sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
            "Frontal Epilepsy": df_126_frontal.apply(
                lambda row:  sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
            "Temporal Epilepsy": df_126_temporal.apply(
                lambda row:  sum((row - controls_126_mean) / controls_126_std >= 2)
                ,
                axis=1,
            ).reset_index(drop=True),
        }
    )

    # for each row in result_mat_lin_age get % of areas above 2 sd of controls
    df_2sd = pd.DataFrame(
        {
            "Controls": controls_126.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std,
                axis=1,
            ).reset_index(drop=True).mean(),
            "Epilepsy": df_126.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(),
            "Focal Epilepsy": df_126_focal.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(),
            "Generalized Epilepsy": df_126_general.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std,
                axis=1,
            ).reset_index(drop=True).mean(),
            "Frontal Epilepsy": df_126_frontal.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(),
            "Temporal Epilepsy": df_126_temporal.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(),
        }
    )
    df_2sd_t = pd.DataFrame(
        {
            "Controls": controls_126.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(axis=1),
            "Epilepsy": df_126.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(axis=1),
            "Focal Epilepsy": df_126_focal.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(axis=1),

            "Generalized Epilepsy": df_126_general.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(axis=1),
            "Temporal Epilepsy": df_126_temporal.apply(
                lambda row:
                 (row - controls_126_mean) / controls_126_std ,
                axis=1,
            ).reset_index(drop=True).mean(axis=1),
        }
    )
    df2 = pd.concat([df_BBB_percent, df_2sd], axis=1)
    # num_regions_to_plots(clinical_data_df)
    # df to csv
    # df_2sd_t.to_csv("figures/df_126_areas_mean.csv", index=False)
    # df_BBB_percent.to_csv("figures/df_BBB_percent.csv", index=False)
    # df_2bbbd.to_csv("figures/bbbd_126_areas.csv", index=False)
    # df_2sd.to_csv("figures/zscore126.csv", index=False)
    # medicine_df(clinical_data_df,result_mat_lin_age,df_126,controls_126_mean,controls_126_std,df_2bbbd)
    # plots(df,df_2sd)
    def matching_bbb_clinical(Epilepsy_clinical_data):
        Epilepsy_clinical_data = Epilepsy_clinical_data[Epilepsy_clinical_data['Focal/General'].notna()]
        df = pd.DataFrame()
        di = {'frontal': 0, 'temporal': 0, 'generalized': 0,'occipital':0,'parietal':0}
        di_sum = {'frontal': 0, 'temporal': 0, 'generalized': 0,'occipital':0,'parietal':0}
        di_contra = {'frontal': 0, 'temporal': 0, 'generalized': 0,'occipital':0,'parietal':0}
        for index,row in Epilepsy_clinical_data.iterrows():
            code =row['code']
            ftype = row['Focal Type']
            areas = result_mat_lin_age126[result_mat_lin_age126["'ID'"]==code.strip("'")].drop(columns=["'ID'"])
            # strip "'" from areas index
            areas = areas.rename(columns=lambda x: x.strip("'")).T
            areas_above_controls = (areas.iloc[:,0] - controls_126_mean) / controls_126_std
            if ftype != ftype:
                continue
                if row['Focal/General'].find('G') != -1:
                    di_sum['generalized'] += 1
                    # if more than 50% of areas_above_controls are above 2 sd of controls
                    if sum(areas_above_controls > 2)/len(areas_above_controls) > 0.5:
                        di['generalized'] += 1
                    df = pd.concat([df,areas_above_controls],axis=1)
                    continue
            if ftype.find('R') != -1:
                side = 'Right'
            elif ftype.find('L') != -1:
                side = 'Left'
            else:
                continue
                side = ' '
            # run over ftype letters TFOP
            for letter in re.findall("[TFOP]", ftype):
                # if ftype contains T
                if letter == 'T':
                    # find index in areas with "temporal"
                    lobe = 'temporal'
                elif letter == 'F':
                    # find index in areas with "frontal"
                    lobe = 'frontal'
                elif letter == 'O':
                    # find index in areas with "occipital"
                    lobe = 'occipital'
                elif letter == 'P':
                    # find index in areas with "parietal"
                    lobe = 'parietal'
                q_areas = areas_above_controls[areas_above_controls.index.str.contains(lobe)]
                q_areas = q_areas[q_areas.index.str.contains(side)]
                # if sum(q_areas > 2)/len(q_areas) > 0.5:
                #     di[lobe] += 1
                di_sum[lobe] += 1
                di[lobe] += 1 if any(q_areas >= 2) else 0
                if side != ' ':
                    other_side = 'Left' if side == 'Right' else 'Right'
                    q_areas = areas_above_controls[areas_above_controls.index.str.contains(lobe)]
                    q_areas = q_areas[q_areas.index.str.contains(other_side)]
                    di_contra[lobe] += 1 if any(q_areas >= 2) else 0
                df = pd.concat([df,q_areas],axis=1)
        print(sum(di.values())/sum(di_sum.values()))
        di = {k: 100*v / di_sum[k] if di_sum[k] != 0 else 0 for k, v in di.items()}
        di_contra = {k: 100*v / di_sum[k] if di_sum[k] != 0 else 0 for k, v in di_contra.items()}
        df.to_csv("figures/matching_bbb_clinical.csv")
    # Example usage
    # plot_shap_feature_importance(clinical_data_df, '# regions with BBBD')
    matching_bbb_clinical(Epilepsy_clinical_data)
    return df_BBB_percent, df_2sd

import re
def tonic_clonic(Epilepsy_clinical_data):
    Epilepsy_clinical_data2 = Epilepsy_clinical_data.iloc[:Epilepsy_clinical_data["Seizures type"].dropna().str.contains("TCS").shape[0]]
    df = pd.DataFrame(
        {
            "TCS": Epilepsy_clinical_data2[Epilepsy_clinical_data2["Seizures type"].dropna().str.contains("TCS")]['# regions with BBBD'].reset_index(drop=True),
            "Non-TCS": Epilepsy_clinical_data2[~Epilepsy_clinical_data2["Seizures type"].dropna().str.contains("TCS")]['# regions with BBBD'].reset_index(drop=True),
        }
    )
    create_scientific_boxplot(df,y_label="#, Regions",filename = 'Tonic_clonic')

def facility(clinical_data_df,df_2bbbd):
    clinical_data_df2 = clinical_data_df[clinical_data_df['Focal/General'].notna()]
    df = pd.DataFrame(
        {
        'BGU': clinical_data_df2[clinical_data_df2["Facility"].dropna().str.contains("BGU")]['# regions with BBBD'].reset_index(drop=True),
        'DAL': clinical_data_df2[clinical_data_df2["Facility"].dropna().str.contains("DAL")]['# regions with BBBD'].reset_index(drop=True),
        'TRI': clinical_data_df2[clinical_data_df2["Facility"].dropna().str.contains("TRI")]['# regions with BBBD'].reset_index(drop=True),
        'UCL': clinical_data_df2[clinical_data_df2["Facility"].dropna().str.contains("UCL")]['# regions with BBBD'].reset_index(drop=True),
        }
    )
    create_scientific_boxplot(df,y_label="#, Regions",filename = 'Facility')
    df = pd.DataFrame(
        {
            'BGU': df_2bbbd['Controls'].iloc[:43].reset_index(drop=True),
            'UCL': df_2bbbd['Controls'].iloc[43:50].reset_index(drop=True),
            'DAL': df_2bbbd['Controls'].iloc[50:].reset_index(drop=True),
        }
    )
    create_scientific_boxplot(df,y_label="#, Regions",filename = 'Facility_controls')

def num_regions_to_plots(clinical_data_df):

    clinical_data_df.columns
    # clinical_data_df['gender'] drop '
    clinical_data_df['gender'] = clinical_data_df['gender'].str.strip("'")
    # clinical_data_df['gender'], change F to 1 and M to 0
    clinical_data_df['gender'] = clinical_data_df['gender'].map({'F': 1, 'M': 0})
    # high_bbb_group bigger than median
    high_bbb_group = clinical_data_df[clinical_data_df["BBB%"] > clinical_data_df["BBB%"].median()]
    # low_bbb_group smaller than median
    low_bbb_group = clinical_data_df[clinical_data_df["BBB%"] <= clinical_data_df["BBB%"].median()]
    cols_to_run = ['Year of epilepsy','Age of oneset','age','Seizure Frequency (/m)','gender']
    for col in cols_to_run:
        lowbbb= pd.to_numeric(low_bbb_group[col], errors='coerce').dropna().reset_index(drop=True)
        highbbb = pd.to_numeric(high_bbb_group[col], errors='coerce').dropna().reset_index(drop=True)
        df = pd.concat([highbbb, lowbbb], axis=1)
        # rename columns
        df.columns = ['High BBBD', 'Low BBBD']
        # boxplot Year of epilepsy
        if col == 'Seizure Frequency (/m)':
            col = 'Seizure Frequency'
        create_scientific_boxplot(df,y_label=col,palette=sns.color_palette(["#000000", "#FF0000"]),filename = col)

import shap
import xgboost as xgb
import matplotlib.pyplot as plt

def plot_shap_feature_importance(clinical_data_df, target_column):
    clinical_data_df.columns
    features = ['Focal/General', 'Focal Type', 'age', 'gender', 
                 'Epilepsy type', 'Seizures type', 'Medications', 'Number of medications', 'Divalproex', 'Tegretol',
                   'Levetiracetam', 'Lamotrigine', 'Divalproex.1', 'Lacosamide', 'Family History', 'Age of oneset', 'Year of epilepsy',
                     'Seizure Frequency (/m)', 'Number of medication','Lesion','EEG']
    df2 = clinical_data_df[features]
    # Splitting the data into features and target
    # Select non-numeric columns
    non_numeric_columns = df2.select_dtypes(include=['object', 'category']).columns

    # Apply one-hot encoding to non-numeric columns
    X = pd.get_dummies(df2, columns=non_numeric_columns, drop_first=True)
    # true =1 false =0
    X = X.replace({True: 1, False: 0})
    y = clinical_data_df[target_column]
    # y drop na
    y = y.dropna()
    # x same number of rows as y
    X = X.iloc[:len(y)]
    # Train a model (XGBoost in this case)
    model = xgb.XGBRegressor()
    model.fit(X, y)
    # Calculate SHAP values
    explainer = shap.Explainer(model, X)
    shap_values = explainer(X)
    # Plot the SHAP values
    shap.summary_plot(shap_values, X, plot_type="bar")
    return


def bar_plot(df,y_label,palette,filename):
    # df columns to float
    df = df.astype(float)
    # sns bar plot 
    plt.figure(figsize=(10, 6))
    df.plot(kind='bar', color=palette, legend=False)
    plt.ylabel(y_label)
    plt.savefig(f'figures/{filename}.png', dpi=600, bbox_inches='tight',transparent=True, pad_inches=0.2, quality=95)
    plt.close()

def pair_plot(df,df_2sd_t,df_2bbbd):
    # pairplot
    a = df['Epilepsy'].dropna().astype(float)
    b = df_2bbbd['Epilepsy'].dropna().astype(float)
    c = df_2sd_t['Epilepsy'].dropna().astype(float)
    # pairplot with linear regression and text slope
    g = sns.pairplot(
        data=pd.DataFrame(
            {
                "BBBD": b,
                "Zscore": c,
                "BBB%": a,
            }
        ),
        kind="reg",
    )
    # Add text with slope values
    for ax in g.axes.flat:
        for line in ax.lines:
            slope = line.get_slope()
            if slope is not None:
                ax.annotate(f"Slope: {slope:.2f}",
                            xy=(0.05, 0.95),
                            xycoords=ax.transAxes,
                            ha='left', va='top')
            break

def age_gender_ect(clinical_data_df,df_2sd_t,df):
    """
    Regression analysis of brain volume in patients with epilepsy who have a lesion showed 10.77 ± 7.81% of voxels exhibited BBBD, while the average z-score for all regions was 1.27 ± 1.41.
    """
    clinical_data_df["'gender'"] = clinical_data_df["'gender'"].str.strip("'")
    gender_zscore = pd.DataFrame( {'Men': df_2sd_t['Epilepsy'].dropna()[clinical_data_df["'gender'"] == 'M'].reset_index(drop=True),
    'Women': df_2sd_t['Epilepsy'].dropna()[clinical_data_df["'gender'"] == 'F'].reset_index(drop=True),}
    )
    gender_bbb = pd.DataFrame(
        {'Men': df['Epilepsy'].dropna()[clinical_data_df["'gender'"] == 'M'].reset_index(drop=True),
        'Women': df['Epilepsy'].dropna()[clinical_data_df["'gender'"] == 'F'].reset_index(drop=True),}
    )
    gender_bbb['Men'] = pd.to_numeric(gender_bbb['Men'], errors='coerce')
    gender_bbb['Women'] = pd.to_numeric(gender_bbb['Women'], errors='coerce')
    stat, p = mannwhitneyu(gender_bbb['Men'].dropna(), gender_bbb['Women'].dropna())
    gender_zscore['Men'] = pd.to_numeric(gender_zscore['Men'], errors='coerce')
    gender_zscore['Women'] = pd.to_numeric(gender_zscore['Women'], errors='coerce')
    stat, p2 = mannwhitneyu(gender_zscore['Men'].dropna(), gender_zscore['Women'].dropna())
    str_gender = f"""
    Regression analysis of brain volume in men with epilepsy who have a lesion showed {gender_bbb['Men'].mean():.2f} ± {gender_bbb['Men'].std():.2f}% of voxels exhibited BBBD, while the average z-score for all regions was {gender_zscore['Men'].mean():.2f} ± {gender_zscore['Men'].std():.2f}.
    Regression analysis of brain volume in women with epilepsy who have a lesion showed {gender_bbb['Women'].mean():.2f} ± {gender_bbb['Women'].std():.2f}% of voxels exhibited BBBD, while the average z-score for all regions was {gender_zscore['Women'].mean():.2f} ± {gender_zscore['Women'].std():.2f}.
    Statistical comparison demonstrated insignificant differences in men compared to women of BBB% (p-value = {p2:.2f}) as well as in average z-score for all regions (p-value = {p:.2f}).
    """.replace('\n',' ')
    from scipy import stats
    # Assuming 'age' and 'bbb' are your data
    age = clinical_data_df["'age'"]
    bbb = df['Epilepsy'].dropna().astype(float)
    zscore = df_2sd_t['Epilepsy'].dropna().astype(float)
    slope, intercept, r_value, p_value, std_err = stats.linregress(age, bbb)
    slope2, intercept2, r_value2, p_value2, std_err2 = stats.linregress(age, zscore)
    Age_of_onset = clinical_data_df['Age of oneset'].dropna().astype(float)
    slope3, intercept3, r_value3, p_value3, std_err3 = stats.linregress(Age_of_onset, bbb)
    slope33, intercept33, r_value33, p_value33, std_err33 = stats.linregress(Age_of_onset, zscore)
    Year_of_epilepsy = clinical_data_df['Year of epilepsy'].dropna().astype(float)
    slope4, intercept4, r_value4, p_value4, std_err4 = stats.linregress(Year_of_epilepsy, bbb)
    slope44, intercept44, r_value44, p_value44, std_err44 = stats.linregress(Year_of_epilepsy, zscore)
    str_age = f"""
    Statistical comparison demonstrated insignificant differences between age and BBB% (p-value = {p_value:.2f}) as well as and averaged z-score for all regions (p-value = {p_value2:.2f}).
    In addition, statistical comparison demonstrated insignificant differences between age of onset and BBB% (p-value = {p_value3:.2f}) as well as and averaged z-score for all regions (p-value = {p_value33:.2f}).
    Finally, statistical comparison demonstrated insignificant differences between years of epilepsy and BBB% (p-value = {p_value4:.2f}) as well as and averaged z-score for all regions (p-value = {p_value44:.2f}).
    """.replace('\n',' ')


import numpy as np
def most_penetrable_regions(df_2sd):
    region_epilepsy = df_2sd['Epilepsy'].sort_values(ascending=False).iloc[:10].apply(lambda x: round(x, 2))
    region_controls = df_2sd['Controls'].sort_values(ascending=False).iloc[:10]
    region_focal = df_2sd['Focal Epilepsy']
    region_general = df_2sd['Generalized Epilepsy']
    region_focal = merge_regions_sides(region_focal).sort_values(ascending=False)
    region_general = merge_regions_sides(region_general).sort_values(ascending=False)
    i = 1
    while True:
        region_focal_temp = merge_regions_sides(region_focal).sort_values(ascending=False).iloc[:i].apply(lambda x: round(x, 2))
        region_general_temp = merge_regions_sides(region_general).sort_values(ascending=False).iloc[:i].apply(lambda x: round(x, 2))
        indexes = region_focal_temp[region_focal_temp.index.isin(region_general_temp.index)].index
        if len(indexes) >= 3 or i >= min(len(region_focal), len(region_general)):
            break
        i += 1
    region_focal_temp[indexes]
    region_general_temp[indexes]
    str_paper = f"""
    Focal regions with most BBBD in patients with epilepsy are {region_focal_temp[indexes]}.
    """.replace('\n', '')
def merge_regions_sides(df):
    # df remove from 'Left' and 'Right'
    df.index = df.index.str.replace('Left ', '').str.replace('Right ', '')
    # find similar indexes and average them
    df = df.groupby(df.index).max()
    return df

def plots(df_2bbbd,df_2sd,clinical_data_df,df_2sd_t):
    create_scientific_boxplot(df_2bbbd[['Controls','Epilepsy']],y_label="#, Regions",palette=sns.color_palette(["#000000", "#FF0000"]),filename = 'BBB_controls_epilepsy')
    create_scientific_boxplot(df_2sd[['Controls','Epilepsy']],y_label="Zscore",palette=sns.color_palette(["#000000", "#FF0000"]),filename = 'zsocre_controls_epilepsy')
    create_scientific_boxplot(df_2bbbd[['Generalized Epilepsy','Focal Epilepsy','Frontal Epilepsy']],y_label="#, Regions",palette=sns.color_palette(["#FF0000", "#990000", "#660000"]),filename = 'BBB_general_focal_frontal_epilepsy')
    create_scientific_boxplot(df_2sd[['Generalized Epilepsy','Focal Epilepsy','Frontal Epilepsy']],y_label="Zscore",palette=sns.color_palette(["#FF0000", "#990000", "#660000"]),filename = 'zsocre_general_focal_frontal_epilepsy')

def lesion_df(df_2bbbd,clinical_data_df,df_2sd_t):
    df_e = df_2bbbd['Epilepsy'].dropna()
    # Lesion
    dfbbbd = pd.DataFrame(
        {
            "Non-lesional": df_e[clinical_data_df['Lesion'] ==0].reset_index(drop=True),
            "Lesion": df_e[clinical_data_df['Lesion'] !=0].reset_index(drop=True),
        }
    )
    df2 = pd.DataFrame(
        {
            "Non-lesional": df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] ==0].reset_index(drop=True),
            "Lesion": df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] !=0].reset_index(drop=True),
        }
    )
    df_all = pd.concat([dfbbbd, df2], axis=1)
    df_all.to_csv("figures/df_lesion.csv", index=False)
    # create_scientific_boxplot(dfbbbd,y_label="BBBD%",palette=sns.color_palette(["#000000", "#FF0000"]),filename = 'BBB_lesion')
    # create_scientific_boxplot(df2,y_label="Zscore",palette=sns.color_palette(["#000000", "#FF0000"]),filename = 'zsocre_lesion')
    f""" {df2['Lesion'].mean():.2f} ± {df2['Lesion'].std():.2f} vs {df2['Non-lesional'].mean():.2f} ± {df2['Non-lesional'].std():.2f}, p={mannwhitneyu(df2['Lesion'].dropna(),df2['Non-lesional'].dropna())[1]:.2f}"""
    f""" {dfbbbd['Lesion'].mean():.2f} ± {dfbbbd['Lesion'].std():.2f} vs {dfbbbd['Non-lesional'].mean():.2f} ± {dfbbbd['Non-lesional'].std():.2f}, p={mannwhitneyu(dfbbbd['Lesion'].dropna(),dfbbbd['Non-lesional'].dropna())[1]:.2f}"""
    str_paper = f"""
    Regression analysis of brain volume in patients with epilepsy who have a lesion showed {df_e[clinical_data_df['Lesion'] !=0].mean():.2f}  ± {df_e[clinical_data_df['Lesion'] !=0].std():.2f}% of voxels exhibited BBBD,
    while the average z-score for all regions was {df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] !=0].mean():.2f} ± {df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] !=0].std():.2f}.
    In addition, regression analysis of brain volume in patients with epilepsy who do not have a lesion showed {df_e[clinical_data_df['Lesion'] ==0].mean():.2f}  ± {df_e[clinical_data_df['Lesion'] ==0].std():.2f}% of voxels exhibited BBBD,
    while the average z-score for all regions was {df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] ==0].mean():.2f} ± {df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] ==0].std():.2f}.
    Statistical comparison demonstrated insignificant differences between lesional and non-lesional patients in both percent of regions with BBBD (p={mannwhitneyu(df_e[clinical_data_df['Lesion'] !=0],df_e[clinical_data_df['Lesion'] ==0])[1]:.2f}),
    and z-score (p={mannwhitneyu(df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] !=0],df_2sd_t['Epilepsy'].dropna()[clinical_data_df['Lesion'] ==0])[1]:.2f}).
    """.replace('\n',' ')
    dfbbbd.to_csv("figures/df_lesion.csv", index=False)


def medicine_df(clinical_data_df,result_mat_lin_age,df_126,controls_126_mean,controls_126_std,df_2bbbd,df_2sd_t):
    medications = ['Phenytoin', 'Divalproex', 'Lamotrigine', 'Levetiracetam', 'Ethosuximide', 'Perampanel', 'Brivaracetam', 'Gabapentin', 'Trileptin', 'Eslicarbazepine', 'Clobazam', 'Lacosamide', 'Zonisamide', 'Eslicarbazepine', 'Topiramate','Frisium','Tegretol','Carbamazepine','OXCARBAZEPINE','Lorazepam']
    # medications remove duplicates
    medications =set(medications)
    med_series = clinical_data_df["Medications"]
    # make list of how many medications each patient has
    med_list = []
    for meds in med_series:
        if meds != meds:
            med_list.append(0)
            continue
        # meds remove duplicates of words
        meds = ' '.join(set(meds.split(' ')))
        # input("Press enter to continue to the next iteration...")
        if pd.isna(meds):
            med_list.append(0)
        else:
            # count how many items from list medications appear in meds
            med_list.append(len([x for x in medications if x in str(meds)]))
    pd.DataFrame(med_list).to_csv("figures/df_medications.csv", index=False)
    # create df
    pd.DataFrame({"1": clinical_data_df.loc[[x<=1 for x in med_list],:]['Seizure Frequency (/m)'].dropna().reset_index(drop=True),
                       "2": clinical_data_df.loc[[x==2 for x in med_list],:]['Seizure Frequency (/m)'].dropna().reset_index(drop=True),
                       "3+": clinical_data_df.loc[[x>2 for x in med_list],:]['Seizure Frequency (/m)'].dropna().reset_index(drop=True)
                       }).to_csv("figures/df_seizure_freq.csv", index=False)
    df_bbb = pd.DataFrame({"1": result_mat_lin_age.loc[[x<=1 for x in med_list],:]['y_target'].reset_index(drop=True)
                       , "2": result_mat_lin_age.loc[[x==2 for x in med_list],:]['y_target'].reset_index(drop=True)
                       , "3+": result_mat_lin_age.loc[[x>2 for x in med_list],:]['y_target'].reset_index(drop=True)
                       })
    df_2bbbd_e = df_2bbbd['Epilepsy'].dropna()
    df_126_med = pd.DataFrame({"1": df_2bbbd_e.loc[[x<=1 for x in med_list]].reset_index(drop=True)
                       , "2": df_2bbbd_e.loc[[x==2 for x in med_list]].reset_index(drop=True)
                       , "3+": df_2bbbd_e.loc[[x>2 for x in med_list]].reset_index(drop=True)
                       })
    zscore_med = pd.DataFrame({"1": df_2sd_t['Epilepsy'].dropna().loc[[x<=1 for x in med_list]].reset_index(drop=True),
                       "2": df_2sd_t['Epilepsy'].dropna().loc[[x==2 for x in med_list]].reset_index(drop=True),
                       "3+": df_2sd_t['Epilepsy'].dropna().loc[[x>2 for x in med_list]].reset_index(drop=True)
                       })
    df_all = pd.concat([df_bbb,df_126_med],axis=1)
    df_all.to_csv("figures/df_medicines.csv", index=False)
    # create_scientific_boxplot(df_bbb,y_label="BBB%",palette=sns.color_palette(["#FF0000", "#990000", "#660000"]),filename = 'BBB_medications')
    # create_scientific_boxplot(df_126_med,y_label="%",palette=sns.color_palette(["#FF0000", "#990000", "#660000"]),filename = 'zsocre_medications')
    f""" {df_126_med['1'].mean():.2f} ± {df_126_med['1'].std():.2f}, {df_126_med['2'].mean():.2f} ± {df_126_med['2'].std():.2f} and {df_126_med['3+'].mean():.2f} ± {df_126_med['3+'].std():.2f} """
    f""" {zscore_med['1'].mean():.2f} ± {zscore_med['1'].std():.2f}, {zscore_med['2'].mean():.2f} ± {zscore_med['2'].std():.2f} and {zscore_med['3+'].mean():.2f} ± {zscore_med['3+'].std():.2f} """
    from scipy.stats import f_oneway
    # Assuming zscore_med is your DataFrame
    f_val, p_val = f_oneway(zscore_med['1'].dropna(), zscore_med['2'].dropna(), zscore_med['3+'].dropna())
    # run on df_126_med
    # Assuming df_126_med is your DataFrame
    f_val2, p_val2 = f_oneway(df_126_med['1'].dropna(), df_126_med['2'].dropna(), df_126_med['3+'].dropna())
    return med_list
    median_lin = result_mat_lin_age['y_target'].median()
    # df above 10 y_target at result_mat_lin_age
    df_above_10 = pd.DataFrame(
        {
            'High BBB%': result_mat_lin_age.loc[[x > median_lin for x in result_mat_lin_age['y_target']], :][
                'y_target'
            ].reset_index(drop=True)
            ,
            'Low BBB%': result_mat_lin_age.loc[[x <= median_lin for x in result_mat_lin_age['y_target']], :][
                'y_target'
            ].reset_index(drop=True)
        }
    )
    f""" {df_above_10['High BBB%'].mean():.2f} ± {df_above_10['High BBB%'].std():.2f}, {df_above_10['Low BBB%'].mean():.2f} ± {df_above_10['Low BBB%'].std():.2f} """

import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import plotting
from scipy.stats import mannwhitneyu


def create_scientific_xyplot(df_bbb, y_label="Value", palette="Set1", filename = 'a',x_label="Medicine Count"):
    # Set the style
    sns.set(style="white")
    # Create the scatterplot
    plt.figure(figsize=(8, 6))
    font_size = 18
    df_melted = df_bbb.melt(var_name=x_label, value_name=y_label)
    ax = sns.scatterplot(x=x_label, y=y_label, data=df_melted, palette=palette)
    # Customize the plot
    ax.set_ylabel(y_label, fontsize=font_size)
    ax.tick_params(labelsize=font_size)
    plt.xticks(fontsize=font_size)
    plt.yticks(fontsize=font_size)
    # remove x axis label
    ax.set_xlabel(None)
    # New code to remove the legend title
    legend = ax.get_legend()
    legend.set_title(None)
    # legend text size
    for text in legend.get_texts():
        text.set_fontsize(int(font_size*0.8))
    # Display the plot
    plt.show(block=False)
    # Save the plot as an image (e.g., in PDF or PNG format)
    plt.savefig(f"figures/{filename}.png", dpi=900, bbox_inches="tight")
    pass

def create_scientific_boxplot(df_bbb,y_label="Value",palette="Set1",filename = 'a'):
    from scipy.stats import mannwhitneyu
    # Set the style
    sns.set(style="white")
    # Create the boxplot
    plt.figure(figsize=(8, 6))
    font_size = 18
    # Perform the Mann-Whitney U test between each pair of columns
    for i in range(len(df_bbb.columns)):
        for j in range(i+1, len(df_bbb.columns)):
            # Ignore NaNs
            data1 = df_bbb.iloc[:, i].dropna()
            data2 = df_bbb.iloc[:, j].dropna()
            # coerced to float
            data1 = data1.astype(float)
            data2 = data2.astype(float)
            stat, p = mannwhitneyu(data1, data2)
            print(f"{df_bbb.columns[i]} vs. {df_bbb.columns[j]}: U-statistic={stat:.2f}, p-value={p:.4f}")
            if p<0.05:
                # Add a line between the two boxes
                ax = plt.gca()
                y_height = df_bbb.max().max() * 1.02
                # get the color of ax at point (i, y_height)
                try:
                    # get number of  ax.lines
                    num_lines = len(ax.lines)
                    y_height *= (1 + (.03 * num_lines))
                except:
                    y_height = y_height
                ax.plot([i, j], [y_height, y_height], color='gray', linestyle='-', linewidth=1)
                # write number of * based on p value
                astrixs = plotting.convert_pvalue_to_asterisks(p)
                # Write above the middle of the line the asterisks
                ax.text((i + j) / 2, y_height, astrixs, ha='center', va='center',fontsize=font_size, color='gray')
    df_melted = df_bbb.melt(var_name='Medicine Count', value_name='Value')
    ax = sns.boxplot(data=df_bbb, palette=palette, width=0.7, boxprops=dict(facecolor=(1, 1, 1, 0)), showfliers=False)
    ax = sns.swarmplot(x='Medicine Count', y='Value', data=df_melted, hue='Medicine Count', palette=palette, facecolor='none')
    # Customize the plot
    # ax.set_xlabel("Number of Medicines", fontsize=16)
    ax.set_ylabel(y_label, fontsize=font_size)
    # plt.title(f"{y_label} vs Medicines", fontsize=font_size)
    # Customize font sizes for better readability
    ax.tick_params(labelsize=font_size)
    plt.xticks(fontsize=font_size)
    plt.yticks(fontsize=font_size)
    # remove x axis label
    ax.set_xlabel(None)
    # New code to remove the legend title
    legend = ax.get_legend()
    legend.set_title(None)
    # legend text size
    for text in legend.get_texts():
        text.set_fontsize(int(font_size*0.8))
    # Display the plot
    plt.show(block=False)
    # Save the plot as an image (e.g., in PDF or PNG format)
    plt.savefig(f"figures/{filename}.png", dpi=900, bbox_inches="tight")
    pass


def results_paper_dyn():
    # import mannwhitneyu
    from scipy.stats import mannwhitneyu
    import math
    df, df_2sd = get_results()
    mat = get_mat()
    mat_lin_control = mat["result_mat_lin_age_control"]
    mat_tofts_control = mat["result_mat_tofts_age_control"]
    mat_lin = mat["result_mat_lin_age"]
    mat_tofts = mat["result_mat_tofts_age"]
    focal_pat_mat_lin = mat["focal_pat_mat_lin"]
    focal_pat_mat_tofts = mat["focal_pat_mat_tofts"]
    general_pat_mat_lin = mat["general_pat_mat_lin"]
    general_pat_mat_tofts = mat["general_pat_mat_tofts"]
    controls_avg_regions = ((mat_lin_control - mat_lin_control.mean(axis=0)) / mat_lin_control.std(axis=0) >= 2).sum(axis=1)
    mat_lin_avg_regions = ((mat_lin - mat_lin_control.mean(axis=0)) / mat_lin_control.std(axis=0) >= 2).sum(axis=1)
    focal_pat_mat_lin_avg_regions = ((focal_pat_mat_lin - mat_lin_control.mean(axis=0)) / mat_lin_control.std(axis=0) >= 2).sum(axis=1)
    general_pat_mat_lin_avg_regions = ((general_pat_mat_lin - mat_lin_control.mean(axis=0)) / mat_lin_control.std(axis=0) >= 2).sum(axis=1)
    # mann whitney df['Epilepsy'], df['Controls']
    df["Epilepsy"] = pd.to_numeric(df["Epilepsy"], errors="coerce")
    df["Controls"] = pd.to_numeric(df["Controls"], errors="coerce")
    stats, p = mannwhitneyu(df["Epilepsy"].dropna(), df["Controls"].dropna())
    
    exponent = math.floor(math.log10(abs(p)))
    # mann whitney df['Epilepsy'], df['Controls']
    d1 = pd.to_numeric(mat_lin_avg_regions, errors="coerce")
    d2 = pd.to_numeric(controls_avg_regions, errors="coerce")
    stats, p = mannwhitneyu(d1.dropna(), d2.dropna())
    exponent2 = math.floor(math.log10(abs(p)))
    pass
    di = {}
    di['controls'] = f"""
    Regression analysis of brain volume in healthy controls revealed that {df['Controls'].mean():.2f} ± {df['Controls'].std():.2f}% of voxels exhibited BBBD,
    while the averaged percent of regions with BBBD for all regions was {controls_avg_regions.mean():.2f} ± {controls_avg_regions.std():.2f}%.
    """.replace("\n", " ")
    di["Patients with epilepsy"] = f"""
    Regression analysis of brain volume using the in all patients with epilepsy revealed that {df['Epilepsy'].mean():.2f} ± {df['Epilepsy'].std():.2f}% of voxels exhibited BBBD,
    while the averaged percent of regions with BBBD for all regions was {mat_lin_avg_regions.mean():.2f} ± {mat_lin_avg_regions.std():.2f}%.
    Statistical comparisons demonstrated significant differences BBBD% between groups (p<10^{exponent}) as well as in the averaged percent of regions with BBBD for all regions (p<10^{exponent2}).
    with at least one brain region with BBBD in {100*sum(mat_lin_avg_regions>1)/len(mat_lin_avg_regions)}
    """.replace("\n", " ")
    # mann whitney df['Focal Epilepsy'], df['Controls']
    df["Focal Epilepsy"] = pd.to_numeric(df["Focal Epilepsy"], errors="coerce")
    df["Controls"] = pd.to_numeric(df["Controls"], errors="coerce")
    stats, p = mannwhitneyu(df["Focal Epilepsy"].dropna(), df["Controls"].dropna())
    exponent = math.floor(math.log10(abs(p)))
    # mann whitney df['Epilepsy'], df['Controls']
    d1 = pd.to_numeric(focal_pat_mat_lin_avg_regions, errors="coerce")
    d2 = pd.to_numeric(general_pat_mat_lin_avg_regions, errors="coerce")
    stats, p = mannwhitneyu(d1.dropna(), d2.dropna())
    exponent2 = math.floor(math.log10(abs(p)))
    # mann whitney df['Generalized Epilepsy'], df['Controls']
    df["Generalized Epilepsy"] = pd.to_numeric(
        df["Generalized Epilepsy"], errors="coerce"
    )
    df["Controls"] = pd.to_numeric(df["Controls"], errors="coerce")
    stats, p = mannwhitneyu(
        df["Generalized Epilepsy"].dropna(), df["Controls"].dropna()
    )
    exponent3 = math.floor(math.log10(abs(p)))
    # mann whitney df['Epilepsy'], df['Controls']
    d1 = pd.to_numeric(general_pat_mat_lin_avg_regions, errors="coerce")
    d2 = pd.to_numeric(controls_avg_regions, errors="coerce")
    stats, p = mannwhitneyu(d1.dropna(), d2.dropna())
    exponent4 = math.floor(math.log10(abs(p)))
    di["focal_generalized"] = f"""
    Regression analysis of brain volume using the in focal epilepsy revealed that {df['Focal Epilepsy'].mean():.2f} ± {df['Focal Epilepsy'].std():.2f}% of voxels exhibited BBBD,
    while {focal_pat_mat_lin_avg_regions.mean():.2f} ± {focal_pat_mat_lin_avg_regions.std():.2f}.
    For generalized epilepsy, the regression analysis revealed that {df['Generalized Epilepsy'].mean():.2f} ± {df['Generalized Epilepsy'].std():.2f}% of voxels exhibited BBBD,
    while {general_pat_mat_lin_avg_regions.mean():.2f} ± {general_pat_mat_lin_avg_regions.std():.2f}.
    Statistical comparisons demonstrated significant differences in focal epilepsy compares to controls (p<10^{exponent}) as well as in the average z-score for all regions (p<10^{exponent2}).
    Statistical comparisons demonstrated significant differences in Generalized epilepsy compares to controls (p<10^{exponent3}) as well as in the average z-score for all regions (p<10^{exponent4}).
    """.replace(
        "\n", " "
    )
    # Assuming df is your DataFrame
    df_dict = mat_lin_avg_regions.sort_values(ascending=False)[:20].to_dict()
    # Convert the dictionary to a string with the format index:val,index2:val2,index3,val3
    df_str = ','.join(f' {k}: {round(v,2)}' for k, v in df_dict.items())
    di['fig2.1'] = f"""
    The 20 regions with the most averaged z-score across all PWE are:
    {df_str}
    """.replace("\n", " ")
    # Assuming df is your DataFrame
    focal_dict = focal_pat_mat_lin_avg_regions.dropna().sort_values(ascending=False).to_dict()
    generalized_dict = general_pat_mat_lin_avg_regions.dropna().sort_values(ascending=False).to_dict()
    # Assuming focal_dict and generalized_dict are your dictionaries
    result_dict = {k: (focal_dict[k], generalized_dict[k]) for k in focal_dict}
    # Convert the dictionary to a string with the format index:val_focal while val_generalized,index2:val_focal2 while val_generalized2
    result_str = ','.join(f' {k}: {round(v[0],2)}|{round(v[1],2)}' for k, v in result_dict.items())
    di['fig2.2'] = f"""
    The averaged z-score for focal and generalized PWE is:
    {result_str}
    """.replace("\n", " ")
    di[
        "focal_generalized"
    ] = f"""
    The regions with the averaged z-score across all PWE above 2 are:
    """
    table_focal_generalized = pd.DataFrame({"Generalized": general_pat_mat_lin_avg_regions, "Focal": focal_pat_mat_lin_avg_regions})
    # sort generalized
    table_focal_generalized.sort_values(by=['Generalized'], ascending=False, inplace=True)
    # to csv
    table_focal_generalized.to_csv("figures/table_focal_generalized.csv")
    pass


def get_126_areas(clinical_data_df, result_mat_lin_age):
    df_126 = pd.DataFrame()
    codes = clinical_data_df["code"]
    # remove nan from codes
    codes = codes.dropna()
    # codes strip "'"
    codes = codes.str.strip("'")
    # run over result_mat_lin_age
    for index, row in result_mat_lin_age.iterrows():
        if "'ID'" in row:
            # get ID if it is in codes
            Id = row["'ID'"].strip("'")
        else:
            Id = row["ID"]
        try:
            # find index in codes that id is contained in
            index_in_codes = codes[[i for i, x in enumerate(codes) if x in Id][0]]
            # change result_mat_lin_age row["'ID'"] to index_in_codes
            row["'ID'"] = index_in_codes
            result_mat_lin_age.loc[index, "'ID'"] = index_in_codes
            # concat row to df_126
            df_126 = pd.concat([df_126, pd.DataFrame(row).T], ignore_index=True)
        except:
            continue
    # strip "'" from df_126 columns
    df_126.columns = df_126.columns.str.strip("'")
    return df_126


def get_mat():
    files_location, excel_file, controls_file, clinical_data = get_folders()

    clinical_data_df = pd.read_excel(clinical_data, sheet_name="All")
    clinical_data_df["code"] = clinical_data_df["code"].str.strip("'")
    nan_code_index = clinical_data_df[clinical_data_df["code"].isna()].index[0]
    # until row 44
    clinical_data_df = clinical_data_df.iloc[:nan_code_index]
    mat = {}
    # read sheet 128_Reagions_Linear
    result_mat_lin_age = pd.read_excel(excel_file, sheet_name="126_Regions_Linear")
    mat["result_mat_lin_age_control"] = pd.read_excel(controls_file, sheet_name="126_Regions_Linear").drop(columns=["'ID'"])
    df_126 = get_126_areas(clinical_data_df, result_mat_lin_age)
    # remove rows from clinical_data_df where code is not in result_mat_lin_age
    clinical_data_df = clinical_data_df[clinical_data_df["code"].isin(df_126["ID"])].reset_index(drop=True)
    mat["result_mat_lin_age"] = df_126.drop(columns=["ID"])
    # mat['focal_pat_mat_lin'] = result_mat_lin_age where clinical_data_df['Focal/General'] == 'F'
    mat["focal_pat_mat_lin"] = df_126[clinical_data_df["Focal/General"] == "F"].drop(columns=["ID"])
    mat["general_pat_mat_lin"] = df_126[clinical_data_df["Focal/General"] == "G"].drop(columns=["ID"])
    mat["result_mat_tofts_age_control"] = mat["result_mat_lin_age_control"]
    mat["result_mat_tofts_age"] = mat["result_mat_lin_age"]
    mat["focal_pat_mat_tofts"] = mat["focal_pat_mat_lin"]
    mat["general_pat_mat_tofts"] = mat["general_pat_mat_lin"]
    mat["areas_names"] = [
        x.strip("'") for x in list(mat["result_mat_lin_age"].columns.values)
    ]
    # f_areas = areas in mat["areas_names"]  which contain "frontal"
    f_areas_indexes = [i for i, x in enumerate(mat["areas_names"]) if "frontal" in x]
    t_areas_indexes = [i for i, x in enumerate(mat["areas_names"]) if "temporal" in x]
    mat["fronal_epilepsy_f_areas_lin"] = df_126[
        clinical_data_df["Focal Type"].str.contains("F", na=False)
    ].drop(columns=["ID"])
    mat["fronal_epilepsy_f_areas_tofts"] = mat["fronal_epilepsy_f_areas_lin"]
    mat["rest_epilepsy_f_areas_lin"] = df_126[
        ~clinical_data_df["Focal Type"].str.contains("F", na=False)
    ].drop(columns=["ID"])
    mat["rest_epilepsy_f_areas_tofts"] = mat["rest_epilepsy_f_areas_lin"]
    mat["controls_f_areas_lin"] = mat["result_mat_lin_age_control"]
    mat["controls_f_areas_tofts"] = mat["controls_f_areas_lin"]
    mat["temporal_epilepsy_t_areas_lin"] = df_126[
        clinical_data_df["Focal Type"].str.contains("T", na=False)
    ].drop(columns=["ID"])
    mat["temporal_epilepsy_t_areas_tofts"] = mat["temporal_epilepsy_t_areas_lin"]
    mat["rest_epilepsy_t_areas_lin"] = df_126[
        ~clinical_data_df["Focal Type"].str.contains("T", na=False)
    ].drop(columns=["ID"])
    mat["rest_epilepsy_t_areas_tofts"] = mat["rest_epilepsy_t_areas_lin"]
    mat["controls_t_areas_lin"] = mat["result_mat_lin_age_control"]
    mat["controls_t_areas_lin"].columns = mat["controls_t_areas_lin"].columns.str.strip(
        "'"
    )
    mat["controls_t_areas_tofts"] = mat["controls_t_areas_lin"]
    # create df
    return mat
    # print codes which are in clinical_data_df but not in clinical_data_df2
    print(
        clinical_data_df[
            clinical_data_df["code"].isin(clinical_data_df2["code"]) == False
        ]["code"]
    )

import os


def renmaes():
    path = r"D:\Dropbox (BGU DAL BBB Group)\Nir_Epilepsy_Controls\Epilepsy\DICOMS\1489752_DICOM"
    folders = [f for f in os.listdir(path) if os.path.isdir(os.path.join(path, f))]
    print(folders)
    # path2 = r"D:\Dropbox (BGU DAL BBB Group)\Nir_Epilepsy_Controls\Epilepsy\DICOMS\1496759_DICOM"
    path2 = r"D:\Dropbox (BGU DAL BBB Group)\Nir_Epilepsy_Controls\Epilepsy\DICOMS\1493706_DICOM"
    folders2 = [f for f in os.listdir(path2) if os.path.isdir(os.path.join(path2, f))]
    # run over folders2
    for folder in folders2:
        if any(folder in f for f in folders):
            folder_from_folders = [f for f in folders if folder in f][0]
            # rename folder to folder_from_folders
            os.rename(
                os.path.join(path2, folder), os.path.join(path2, folder_from_folders)
            )
            print("found")


def merge_xlsx(path):
    import pandas as pd
    import numpy as np
    import os
    import glob
    import re
    from openpyxl import load_workbook
    # Read excel files from D:\Dropbox (BGU DAL BBB Group)\Nir_Epilepsy_Controls\Epilepsy\Analyse\Excel_results
    all_files = glob.glob(os.path.join(path, "*.xlsx"))
    # read get the names of sheet from all_files[0]
    excel_file = pd.ExcelFile(all_files[0])
    # get the names of sheets
    sheet_names = excel_file.sheet_names
    # run over all_files
    for f in all_files:
        # run over sheet_names
        for sheet in sheet_names:
            # get first row of sheet
            columns_sheet = pd.read_excel(f, sheet_name=sheet, nrows=1).columns
            # if there is a number in columns_sheet
            if any(isinstance(x, float) for x in columns_sheet):
                # read excel with openpyxl
                wb = load_workbook(f)
                shee = wb[sheet]
                # insert row above first
                shee.insert_rows(1)
                # write the first row of sheet ['ID','Lin']
                shee.cell(1, 1).value = "'ID'"
                shee.cell(1, 2).value = "Lin"
                # save
                wb.save(f)
    excel_file_name = os.path.join(path, "merged.xlsx")
    # create empty excel file
    pd.DataFrame().to_excel(excel_file_name)
    for i in range(len(sheet_names)):
        sheet = re.sub(r"\s+", "_", sheet_names[i])
        # run over sheet_names and merge them
        df_from_each_file = (pd.read_excel(f, sheet_name=sheet) for f in all_files)
        df_merged = pd.concat(df_from_each_file, ignore_index=True)
        # df_merged.to_excel("merged.xlsx") in path
        with pd.ExcelWriter(excel_file_name, engine="openpyxl", mode="a") as writer:
            df_merged.to_excel(writer, sheet_name=sheet, index=False)

def rand():
    import numpy as np
    from scipy.stats import chi2_contingency
    from scipy.stats import ttest_ind

    # Given data
    divalproex = [0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0]
    second_set = [0, 1, 0, 1, 1, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 1, 1, 0, 0, 0]

    # Organize data into a contingency table
    contingency_table = np.array([divalproex, second_set])
    # Convert lists to numpy arrays
    contingency_table = np.array([np.array(xi) for xi in contingency_table])
    # do ttest
    t_stat, p_value = ttest_ind(divalproex, second_set)


    chi2_stat, p_value, _, _ = chi2_contingency(contingency_table)
    # Perform the chi-squared test
    chi2_stat, p_value, _, _ = chi2_contingency(contingency_table)

    # Set the significance level
    alpha = 0.05

    # Interpret the results
    print(f"Chi-squared statistic: {chi2_stat}")
    print(f"P-value: {p_value}")


if __name__ == "__main__":
    # merge_xlsx(path = r"C:\Nir\BBB\2\BBB\Excel not age gender match\controls")
    # merge_xlsx(path = r"C:\Nir\BBB\2\BBB\Excel not age gender match\epilepsy")
    results_paper_dyn()
    get_results()
    get_mat()
